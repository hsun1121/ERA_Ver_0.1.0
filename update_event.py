from pathlib import Path
import PySimpleGUI as sg
from openpyxl import load_workbook
import openpyxl

current_dir = Path(__file__).parent
EXCEL_FILE = current_dir / 'Saved_Events.xlsx'


def update_event_gui(row_number, column_number):
    # GUI window with options to delete or edit event
    update_event_layout = [
        [sg.Text('What would you like to do: ', size=(30, 2))],
        [sg.Button('Delete Event'), sg.Button('Edit Event Detail'), sg.Exit()]
    ]

    update_event_window = sg.Window("Update Event", update_event_layout, modal=True)
    update_event_control(update_event_window, row_number, column_number)
    return update_event_window


def update_event_control(update_event_window, row_number, column_number):
    while True:
        event, values = update_event_window.read()
        if event == sg.WIN_CLOSED or event == 'Exit':
            break
        # If user presses delete event button, delete the event they clicked on
        if event == 'Delete Event':
            event_data = load_workbook('Saved_events.xlsx')
            user_data = event_data.active
            # Delete clicked row from excel. Index starts at 0 on first event in GUI window,
            # not first row of excel sheet, so add 2 to get excel row #
            user_data.delete_rows(int(row_number) + 2)
            sg.popup('Event Deleted!')
            # Save updated excel file
            event_data.save('Saved_events.xlsx')
            update_event_window.close()

        # If user presses edit event button
        if event == 'Edit Event Detail':
            event_data = openpyxl.load_workbook('Saved_Events.xlsx')
            user_data = event_data['Sheet1']
            # Get row and column number of event detail from GUI. (Index starts at 0 from GUI
            # window events. Add 2 to row and 1 to column to get event index in excel file)
            cell_row = int(row_number) + 2
            cell_column = int(column_number) + 1
            # Get data value of cell clicked on by user
            cell_data = user_data.cell(cell_row, cell_column).value
            edit_event(cell_data, cell_row, cell_column)
            update_event_window.close()


def edit_event(cell_data, cell_row, cell_column):
    # GUI window to update cell data

    if cell_column == 1:
        edit_data_layout = [
            [sg.Text('Please update the EVENT NAME field')],
            [sg.Text('New Event Name:', size=(15, 1)), sg.InputText(key='Update')],

            [sg.Submit(), sg.Button('Clear'), sg.Exit()]
        ]
    elif cell_column == 2:
        edit_data_layout = [
            [sg.Text('Please update the EVENT LOCATION field')],
            [sg.Text('New Location:', size=(15, 1)), sg.InputText(key='Update')],

            [sg.Submit(), sg.Button('Clear'), sg.Exit()]
        ]
    elif cell_column == 3:
        edit_data_layout = [
            [sg.Text('Please update the EVENT MONTH field')],
            [sg.Text('Month:', size=(15, 1)),
             sg.Combo(['January', 'February', 'March', 'April', 'May', 'June', 'July', 'August',
                       'September', 'October', 'November', 'December'], key='Update')],

            [sg.Submit(), sg.Button('Clear'), sg.Exit()]
        ]
    elif cell_column == 4:
        edit_data_layout = [
            [sg.Text('Please update the EVENT DAY field')],
            [sg.Text('Day:', size=(15, 1)),
             sg.Combo(['1', '2', '3', '4', '5', '6', '7', '8', '9', '10', '11', '12', '13', '14',
                       '15', '16', '17', '18', '19', '20', '21', '22', '23', '24', '25', '26',
                       '27', '28', '29', '30', '31'], key='Update')],

            [sg.Submit(), sg.Button('Clear'), sg.Exit()]
        ]
    elif cell_column == 5:
        edit_data_layout = [
            [sg.Text('Please update the EVENT YEAR field')],
            [sg.Text('Year:', size=(15, 1)),
             sg.Combo(['2023', '2024', '2025'], key='Update')],

            [sg.Submit(), sg.Button('Clear'), sg.Exit()]
        ]
    elif cell_column == 6:
        edit_data_layout = [
            [sg.Text('Please update the WISHLIST ITEM 1 field')],
            [sg.Text('New Wishlist Item:', size=(15, 1)), sg.InputText(key='Update')],

            [sg.Submit(), sg.Button('Clear'), sg.Exit()]
        ]
    elif cell_column == 7:
        edit_data_layout = [
            [sg.Text('Please update the WISHLIST ITEM 2 field')],
            [sg.Text('New Wishlist Item:', size=(15, 1)), sg.InputText(key='Update')],

            [sg.Submit(), sg.Button('Clear'), sg.Exit()]
        ]
    elif cell_column == 8:
        edit_data_layout = [
            [sg.Text('Please update the WISHLIST ITEM 3 field')],
            [sg.Text('New Wishlist Item:', size=(15, 1)), sg.InputText(key='Update')],

            [sg.Submit(), sg.Button('Clear'), sg.Exit()]
        ]

    edit_data_window = sg.Window('Update Event', edit_data_layout, modal=True)
    edit_event_control(edit_data_window, cell_data, cell_row, cell_column)
    return edit_data_window


def edit_event_control(edit_data_window, cell_data, cell_row, cell_column):
    while True:
        event, values = edit_data_window.read()
        if event == sg.WIN_CLOSED or event == 'Exit':
            break
        if event == 'Clear':
            for key in values:
                edit_data_window[key]('')
        # If user submits event detail update, change info in excel
        if event == 'Submit':
            event_data = load_workbook('Saved_events.xlsx')
            user_data = event_data.active
            # Assign input update value to variable
            updated_data = values['Update']
            # Replace event detail with updated data
            user_data.cell(cell_row, cell_column).value = updated_data
            # Save updated excel file
            event_data.save('Saved_events.xlsx')
            sg.popup('Update Saved!')
            edit_data_window.close()

    edit_data_window.close()